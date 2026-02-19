import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import io
import os

# --- 1. 디자인 및 페이지 설정 ---
st.set_page_config(page_title="EDC Validation Tool", page_icon="✅", layout="wide")

# 요청하신 로고 가운데 정렬 및 둥근 모서리 제거(border-radius: 0) 적용
st.markdown("""
    <style>
    /* 모든 버튼의 둥근 모서리 */
    .stButton > button, .stDownloadButton > button {
        background-color: #008fd4;
        color: white;
        border: none;
        border-radius: 15px
        font-weight: bold;
    }
    
    /* 입력창 둥근 모서리 */
    .stTextInput > div > div > input {
        border-radius: 15px
    }

    /* 파일 업로드 박스 둥근 모서리 제거 */
    .stFileUploader > section > div {
        border-radius: 0px !important;
    }

    /* 헤더 포인트 색상 */
    h1, h2, h3 {
        color: #008fd4;
    }

    /* 로고 중앙 정렬을 위한 컨테이너 */
    .logo-container {
        display: flex;
        justify-content: center;
        align-items: center;
        padding-bottom: 20px;
    }
    </style>
""", unsafe_allow_html=True)

# --- 2. 로고 및 타이틀 배치 ---

logo_path = "JNPMEDI_original.jpg"  # 실제 로고 파일명

if os.path.exists(logo_path):
    # 로고와 타이틀을 나란히 배치하기 위해 컬럼 사용
    col_logo, col_title = st.columns([2, 4]) # 비율 조정 가능
    with col_logo:
        st.image(logo_path, width=500) # 로고 크기 조절
    with col_title:
        st.title("EDC Validation Auto-Check System")
else:
    # 로고 파일이 없을 경우 기존처럼 이모지로 표시
    st.title("🏥 EDC Validation Auto-Check System")

st.markdown("---")

# --- 3. 사용자 입력 (사이드바 및 메인) ---

# [섹션 1] 버전 정보 입력
st.subheader("✅ 각 문서의 버전 정보")
col1, col2, col3 = st.columns(3)
with col1:
    blank_ver = st.text_input("Blank eCRF Version", value="1.0")
with col2:
    db_spec_ver = st.text_input("DB Spec Version", value="1.0")
with col3:
    annotated_ver = st.text_input("Annotated CRF Version", value="1.0")

# [섹션 2] 파일 업로드
st.subheader("📁 검증 대상 파일 업로드")
col_doc, col_edc = st.columns(2)

with col_doc:
    st.info("**Database Specification 문서**")
    doc_file = st.file_uploader("📂 ⭣⭣**문서 엑셀 파일**을 업로드하세요⭣⭣", type=['xlsx'], key='doc')

with col_edc:
    st.info("**Entry Screen File (CDMS)**")
    edc_file = st.file_uploader("📂 ⭣⭣**EDC 엑셀 파일**을 업로드하세요⭣⭣", type=['xlsx'], key='edc')

# 내부 템플릿 파일 경로 (같은 폴더에 위치한다고 가정)
TEMPLATE_PATH = 'EDC Validation_template.xlsx'

# --- 4. 로직 함수 정의 ---
def get_clean_spec(file, header_row=0):
    """데이터 전처리 및 Key 생성 함수"""
    try:
        df = pd.read_excel(file, sheet_name='DB Specifications', header=header_row, dtype=str)
    except Exception as e:
        st.error(f"파일 읽기 실패: 시트명('DB Specifications')이 정확한지 확인해주세요. ({e})")
        return pd.DataFrame()

    df = df.fillna('')
    df.columns = [str(c).upper().strip() for c in df.columns]
    
    std_cols = ['DOMAIN', 'DOMAIN LABEL', 'PAGE', 'PAGE LABEL', 'VISIT', 
                'ITEM ID', 'ITEM LABEL', 'ITEM SEQ', 'VERSION', 'CODE', 
                'LAYOUT', 'TYPE', 'MAX_LEN', 'MIN_VAL', 'MAX_VAL']
    
    # 명칭 보정
    if 'VER.' in df.columns:
        df = df.rename(columns={'VER.': 'VERSION'})
    
    for col in std_cols:
        if col not in df.columns: df[col] = ""
        # 소수점 .0 제거
        df[col] = df[col].apply(lambda x: x[:-2] if str(x).endswith('.0') else str(x))
    
    # JOIN KEY 생성
    df['JOIN_KEY'] = (df['DOMAIN'] + df['PAGE'] + df['VISIT'] + df['ITEM ID']).str.replace(r'\s+', '', regex=True).str.upper()
    
    return df[std_cols + ['JOIN_KEY']]

def safe_write(ws, r, c, val):
    """병합된 셀을 고려하여 안전하게 값을 쓰는 함수"""
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = val
                return
    cell.value = val

def process_validation(doc_file, edc_file, template_path, ver_info):
    """전체 검증 로직 실행"""
    # 1. 데이터 로드 및 전처리
    df_doc = get_clean_spec(doc_file, header_row=1) # 문서는 2행부터 헤더
    df_edc = get_clean_spec(edc_file, header_row=0) # EDC는 1행부터 헤더

    if df_doc.empty or df_edc.empty:
        return None

    # 2. Merge
    df_merged = pd.merge(df_doc, df_edc, on='JOIN_KEY', how='outer', suffixes=('_Doc', '_EDC'), indicator=True)

    # 3. 템플릿 로드 (내부 파일)
    if not os.path.exists(template_path):
        st.error(f"❌ 템플릿 파일을 찾을 수 없습니다: {template_path}")
        return None
    
    wb = load_workbook(template_path)
    ws = wb['Entry Screen Validation']

    # 4. 문서별 버전 정보 기입
    ws['A2'] = f"Blank eCRF Version: {ver_info['blank']}"
    ws['A3'] = f"Database Specifications Version: {ver_info['db']}"
    ws['A4'] = f"Annotated CRF Version: {ver_info['annotated']}"

    # 5. 비교 로직 수행
    std_cols = ['DOMAIN', 'DOMAIN LABEL', 'PAGE', 'PAGE LABEL', 'VISIT', 
                'ITEM ID', 'ITEM LABEL', 'ITEM SEQ', 'VERSION', 'CODE', 
                'LAYOUT', 'TYPE', 'MAX_LEN', 'MIN_VAL', 'MAX_VAL']

    for i, row in df_merged.iterrows():
        curr_r = 7 + i
        status = row['_merge']
        
        if status == 'left_only':
            for idx, col in enumerate(std_cols):
                safe_write(ws, curr_r, idx + 1, row[f"{col}_Doc"])
                safe_write(ws, curr_r, idx + 16, "")
            res = "EDC 구현 누락"
            
        elif status == 'right_only':
            for idx, col in enumerate(std_cols):
                safe_write(ws, curr_r, idx + 1, "")
                safe_write(ws, curr_r, idx + 16, row[f"{col}_EDC"])
            res = "문서 Spec에 존재하지 않음"
            
        else: # both
            mismatches = []
            for idx, col in enumerate(std_cols):
                d_val = row[f"{col}_Doc"]
                e_val = row[f"{col}_EDC"]
                safe_write(ws, curr_r, idx + 1, d_val)
                safe_write(ws, curr_r, idx + 16, e_val)
                if d_val != e_val:
                    mismatches.append(col)
            
            res = "True" if not mismatches else f"[{', '.join(mismatches)}] 값 불일치"
        
        safe_write(ws, curr_r, 31, res)

    # 6. 결과를 메모리에 저장 (파일 다운로드용)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- 5. 실행 버튼 및 결과 출력 ---
st.markdown("---")
if st.button("🚀 검증 시작 (Start Validation)"):
    if doc_file and edc_file:
        with st.spinner('데이터를 분석하고 결과를 생성 중입니다...'):
            ver_info = {'blank': blank_ver, 'db': db_spec_ver, 'annotated': annotated_ver}
            result_excel = process_validation(doc_file, edc_file, TEMPLATE_PATH, ver_info)
            
            if result_excel:
                st.success("✅ 분석이 완료되었습니다! 아래 버튼을 눌러 결과를 다운로드하세요.")
                st.download_button(
                    label="📥 결과 리포트 다운로드 (.xlsx)",
                    data=result_excel,
                    file_name="EDC_Validation_Result_Final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.warning("⚠️ 두 개의 파일(문서 Spec, EDC Spec)을 모두 업로드해주세요.")