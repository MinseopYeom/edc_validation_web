import streamlit as st
import pandas as pd
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell

# ============================================================
# 1. 페이지 설정
# ============================================================
icon_path = "blue-white.png"
st.set_page_config(page_title="JNPMEDI EDC Validation", page_icon=icon_path, layout="wide")

TEMPLATE_PATH = 'EDC Validation_template.xlsx'

# ============================================================
# [유지보수 포인트] SYS_ 레이아웃 제외 시 포함 예외 목록 (ITEM ID 기준)
# 추후 비교에 포함시켜야 할 ITEM ID가 생기면 이 리스트에 추가하세요.
# ============================================================
SYS_LAYOUT_WHITELIST = [
    "SUBJID",
    # "SITEID",  # 예시: 추후 추가할 경우 이런 식으로 등록
]

st.markdown("""
    <style>
    .stApp { background-color: #F4F7F6; color: #333333; }
    h1, h2, h3, h4, h5, h6, p, span, div, label {
        color: #2c3e50 !important;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div {
        background-color: #ffffff !important; color: #333333 !important;
        border: 1px solid #dcdcdc; border-radius: 8px;
    }
    ul[data-testid="stSelectboxVirtualDropdown"] li {
        color: #333333 !important; background-color: #ffffff !important;
    }
    .stFileUploader, div[data-testid="stExpander"], div[data-testid="stVerticalBlock"] > div {
        background-color: #ffffff; color: #333333 !important;
        border-radius: 10px; padding: 5px;
    }
    .stFileUploader label { font-weight: bold; font-size: 1.1em; }
    .stButton > button, .stDownloadButton > button {
        width: 100%; background-color: #008fd4; color: #ffffff !important;
        font-weight: bold; border: none; padding: 0.6rem; border-radius: 8px;
        transition: all 0.3s ease; box-shadow: 0 2px 4px rgba(0,143,212,0.3);
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background-color: #006fa3; color: #ffffff !important;
        box-shadow: 0 4px 8px rgba(0,111,163,0.4); transform: translateY(-1px);
    }
    .stButton > button:active { transform: translateY(0px); }
    .success-box {
        padding: 15px; background-color: #e3f2fd; color: #0d47a1 !important;
        border-left: 5px solid #008fd4; border-radius: 4px;
        margin-bottom: 15px; font-weight: 600;
    }
    .error-box {
        padding: 15px; background-color: #ffebee; color: #b71c1c !important;
        border-left: 5px solid #d32f2f; border-radius: 4px;
        margin-bottom: 15px; font-weight: 600;
    }
    </style>
""", unsafe_allow_html=True)


# ============================================================
# 2. 공통 유틸 함수
# ============================================================

@st.cache_resource
def load_excel_file(file):
    """파일을 메모리에 로드 (속도 향상)"""
    return pd.ExcelFile(file)


def get_dynamic_preview(excel_file, sheet_name, header_row):
    """사용자가 선택한 행을 헤더로 적용하여 10개 미리보기 생성"""
    try:
        return pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row, nrows=10, dtype=str)
    except Exception:
        return pd.DataFrame()


def check_columns_status(df):
    """필수 컬럼이 식별되는지 진단"""
    if df.empty:
        return False, "데이터 없음", []

    current_cols = [str(c).upper().strip() for c in df.columns]

    rename_map = {
        'VAR NAME': 'ITEM ID', 'VARIABLE NAME': 'ITEM ID', 'VARIABLE': 'ITEM ID',
        'OID': 'ITEM ID', 'ITEMOID': 'ITEM ID', 'QUESTION OID': 'ITEM ID',
        'FORM': 'PAGE', 'FORM OID': 'PAGE', 'FORM NAME': 'PAGE', 'CRF PAGE': 'PAGE',
        'FOLDER': 'VISIT', 'FOLDER OID': 'VISIT', 'EVENT': 'VISIT', 'VISIT NAME': 'VISIT',
        'DATASET': 'DOMAIN', 'LB DOMAIN': 'DOMAIN', 'DOMAIN NAME': 'DOMAIN',
        'VER.': 'VERSION', 'VER': 'VERSION', 'CRF_VERSION': 'VERSION', 'CRF VERSION': 'VERSION',
    }

    mapped_cols = set()
    for col in current_cols:
        if col in rename_map:
            mapped_cols.add(rename_map[col])
        elif col in {'DOMAIN', 'PAGE', 'VISIT', 'ITEM ID'}:
            mapped_cols.add(col)

    required = {'DOMAIN', 'PAGE', 'VISIT', 'ITEM ID'}
    missing = required - mapped_cols

    if not missing:
        return True, "✅ 필수 컬럼 자동 인식 성공!", []
    else:
        return False, f"⚠️ 필수 컬럼 미식별: {', '.join(missing)}", list(missing)


def apply_sys_layout_filter(df, whitelist):
    """
    DB Spec에서 SYS_ 레이아웃 행을 필터링합니다.
    - LAYOUT이 'SYS_'로 시작하면 제외
    - 단 ITEM ID가 whitelist에 있으면 포함 유지
    """
    if 'LAYOUT' not in df.columns:
        return df, pd.DataFrame()

    whitelist_upper = [item.upper().strip() for item in whitelist]
    is_sys = df['LAYOUT'].str.upper().str.startswith('SYS_')
    is_whitelisted = df['ITEM ID'].str.upper().isin(whitelist_upper)
    exclude_mask = is_sys & ~is_whitelisted

    return df[~exclude_mask].reset_index(drop=True), df[exclude_mask].reset_index(drop=True)


def process_data_final(excel_file, sheet_name, header_row):
    """DB Spec 파일을 읽어 표준화된 DataFrame으로 반환"""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row, dtype=str)
        df.columns = [str(c).upper().strip() for c in df.columns]

        rename_map = {
            'VAR NAME': 'ITEM ID', 'VARIABLE NAME': 'ITEM ID', 'VARIABLE': 'ITEM ID',
            'OID': 'ITEM ID', 'ITEMOID': 'ITEM ID',
            'FORM': 'PAGE', 'FORM OID': 'PAGE', 'FORM NAME': 'PAGE', 'CRF PAGE': 'PAGE',
            'FOLDER': 'VISIT', 'FOLDER OID': 'VISIT', 'EVENT': 'VISIT',
            'DATASET': 'DOMAIN', 'LB DOMAIN': 'DOMAIN',
            'VER.': 'VERSION', 'VER': 'VERSION', 'CRF_VERSION': 'VERSION', 'CRF VERSION': 'VERSION',
        }
        df = df.rename(columns=rename_map)

        std_cols = ['DOMAIN', 'DOMAIN LABEL', 'PAGE', 'PAGE LABEL', 'VISIT',
                    'ITEM ID', 'ITEM LABEL', 'ITEM SEQ', 'VERSION', 'CODE',
                    'LAYOUT', 'TYPE', 'MAX_LEN', 'MIN_VAL', 'MAX_VAL']

        for col in std_cols:
            if col not in df.columns:
                df[col] = ""
            df[col] = (df[col].fillna("").astype(str)
                       .apply(lambda x: x.replace('.0', '').strip() if x.endswith('.0') else x.strip()))

        df['JOIN_KEY'] = (df['DOMAIN'] + df['PAGE'] + df['VISIT'] + df['ITEM ID']
                          ).str.replace(r'\s+', '', regex=True).str.upper()

        df = df[df['JOIN_KEY'].str.len() > 1]
        df = df.drop_duplicates(subset=['JOIN_KEY'])
        return df
    except Exception:
        return pd.DataFrame()


# ============================================================
# 3. Data Structure Validation 관련 함수
# ============================================================

def parse_item_id(col_name: str) -> str:
    """
    'ITEMID:LABEL' 형태의 컬럼명에서 ITEM ID 부분만 추출합니다.
    ':' 가 없으면 컬럼명 그대로 반환합니다.
    """
    return col_name.split(':')[0].strip().upper()


def dtype_to_type_str(dtype) -> str:
    """
    pandas dtype을 사람이 읽기 쉬운 Type 문자열로 변환합니다.
    DB Spec의 TYPE 컬럼과 비교하기 위한 참고값입니다.
    """
    dtype_str = str(dtype)
    if 'datetime' in dtype_str:
        return 'datetime'
    elif 'int' in dtype_str:
        return 'integer'
    elif 'float' in dtype_str:
        return 'float'
    else:
        return 'text'


def infer_db_spec_type(col_series: pd.Series) -> str:
    """
    컬럼의 실제 데이터를 분석하여 DB Spec Type 형식으로 반환합니다.

    판정 규칙 (pandas dtype 우선 참조):
    1. datetime64 dtype → NVARCHAR(10)  (YYYY-MM-DD 형식 기준)
       - str 변환 시 19자리가 되는 오추론 방지
    2. int dtype        → INTEGER
    3. float64 dtype    → 실제 값에 소수점 유무로 분기
       a. 소수점 있는 값이 하나라도 존재 → DECIMAL(전체자리수, 소수자리수)
       b. 모든 값이 정수(.0) 형태        → DECIMAL(전체자리수, 0)
          - "INTEGER"가 아닌 DECIMAL(n,0)으로 저장하여 compare_ds_type에서
            DB Spec의 DECIMAL과 자리수 비교가 가능하도록 함
    4. 문자열(object)   → NVARCHAR(최대길이)
       - NCHAR는 사용하지 않고 전부 NVARCHAR로 통일
    5. 유효 데이터 없음 → '' (빈 문자열)

    [수정 이력]
    - NCHAR 판별 로직 제거 → 문자열은 전부 NVARCHAR로 통일
    - datetime64 dtype 처리 추가 (str 변환 시 NCHAR(19) 오추론 방지)
    - float64 소수점 없는 경우: 기존 DECIMAL(n,1) 보장 → DECIMAL(n,0) 으로 변경
      (실제 소수점 여부를 반영하되, DECIMAL 형식 유지로 compare_ds_type 비교 가능)
    """
    dtype_str = str(col_series.dtype)

    # ── 1. datetime64: YYYY-MM-DD 형식 기준 NVARCHAR(10) ─────
    if 'datetime' in dtype_str:
        valid_dates = col_series.dropna()
        if valid_dates.empty:
            return ''
        return 'NVARCHAR(10)'

    # ── null/빈값 제거 후 유효 값만 추출 ─────────────────────
    valid = col_series.dropna()
    valid = valid[valid.apply(lambda x: str(x).strip() not in ('', 'nan', 'NaN', 'NaT'))]

    if valid.empty:
        return ''

    # ── 2. int dtype → INTEGER ────────────────────────────────
    if 'int' in dtype_str:
        return 'INTEGER'

    # ── 3. float64 dtype → DECIMAL 또는 INTEGER ─────────────────
    # 실제 값에서 소수점 유무를 확인하여 분기
    # - 소수점 있는 값이 하나라도 존재 → DECIMAL(전체자리수, 소수자리수)
    # - 모든 값이 정수(.0) 형태         → INTEGER
    #   (null 혼입으로 float64가 된 경우도 실제 값 기준으로 판단)
    #   단, build_dataset_long에서 DS_TYPE=INTEGER이고 DB Spec이 DECIMAL인 경우
    #   자리수 비교를 위해 DECIMAL(n,0)으로 변환하여 저장함 (해당 함수 참조)
    if 'float' in dtype_str:
        max_total_digits   = 0
        max_decimal_digits = 0
        has_decimal_value  = False

        for v in valid:
            v_str = str(v).strip()
            if '.' in v_str:
                integer_part, decimal_part = v_str.split('.', 1)
                dec_part_clean = decimal_part.rstrip('0')
                if dec_part_clean:
                    has_decimal_value = True
                    int_digits   = len(integer_part.lstrip('-').lstrip('0') or '0')
                    dec_digits   = len(dec_part_clean)
                    total_digits = int_digits + dec_digits
                    max_total_digits   = max(max_total_digits,   total_digits)
                    max_decimal_digits = max(max_decimal_digits, dec_digits)
                else:
                    # .0 형태 → 정수 취급
                    int_digits = len(integer_part.lstrip('-').lstrip('0') or '0')
                    max_total_digits = max(max_total_digits, int_digits)
            else:
                int_digits = len(v_str.lstrip('-').lstrip('0') or '0')
                max_total_digits = max(max_total_digits, int_digits)

        if has_decimal_value:
            return f"DECIMAL({max_total_digits},{max_decimal_digits})"
        else:
            return 'INTEGER'

    # ── 4. 문자열(object) → NVARCHAR (NCHAR 사용 안 함) ──────
    # NCHAR/NVARCHAR는 데이터만으로 구분 불가능하므로 NVARCHAR로 통일
    # DB Spec에 NCHAR가 있으면 base type 불일치(False)로 표시 → 사람이 길이만 검토
    str_vals = valid.apply(lambda x: str(x).strip())
    max_len  = int(str_vals.apply(len).max())

    return f"NVARCHAR({max_len})"


def build_dataset_long(dataset_excel: pd.ExcelFile) -> pd.DataFrame:
    """
    CDMS Dataset 엑셀의 모든 도메인 시트를 읽어 Long format DataFrame으로 변환합니다.

    변환 규칙:
    - 시트명 = DOMAIN
    - 컬럼명 'ITEMID:LABEL' → ITEM ID는 ':' 앞 부분만 추출
    - 모든 컬럼을 Item ID로 처리 (제외 없음)
    - 각 Item ID에 대해 실제 데이터를 분석하여 DB Spec Type 형식으로 DS_TYPE 추론
      (DECIMAL(p,s) / NVARCHAR(n) / NCHAR(n) / INTEGER)
    - 값이 실제로 존재하는(non-null) 첫 번째 행의 SUBJID를 참조 대상자로 기록
    - 모든 대상자에게 값이 없는 경우 DS_TYPE = '', DS_SUBJID = '' 으로 기록

    [수정 이력]
    - 기존: found_type = 실제 셀 값 (str) → 주석 처리
    - 변경: found_type = infer_db_spec_type()로 추론한 DB Spec Type 형식 문자열

    Returns:
        DataFrame with columns: [DOMAIN, ITEM ID, DS_TYPE, DS_SUBJID, DS_INT_DIGITS]
        DS_INT_DIGITS: DS_TYPE이 INTEGER일 때 정수 최대 자리수 (그 외 0)
    """
    skip_sheets = {'SUBJECT_INFO'}
    records = []

    for sheet in dataset_excel.sheet_names:
        if sheet.upper() in skip_sheets:
            continue

        domain = sheet.strip().upper()

        try:
            df = pd.read_excel(dataset_excel, sheet_name=sheet)
        except Exception:
            continue

        if df.empty:
            continue

        # SUBJID 컬럼 원본명 찾기 (SUBJID:xxx 형태일 수 있음)
        subjid_col_raw = None
        for c in df.columns:
            if parse_item_id(c) == 'SUBJID':
                subjid_col_raw = c
                break

        # 모든 컬럼을 Item ID로 처리
        for raw_col in df.columns:
            item_id = parse_item_id(raw_col)

            col_series   = df[raw_col]
            found_subjid = ''

            # ── [수정] DS_TYPE: 실제 셀 값 대신 DB Spec Type 형식으로 추론 ──────
            # 기존 코드 (주석 처리):
            # for idx in df.index:
            #     val = col_series.iloc[idx]
            #     if pd.isna(val) or str(val).strip() == '' or str(val).strip().lower() == 'nan':
            #         continue
            #     found_type = str(val).strip()   # ← 실제 셀 값을 그대로 사용하던 부분
            #     if subjid_col_raw is not None:
            #         subj_val = df[subjid_col_raw].iloc[idx]
            #         found_subjid = str(subj_val).strip() if pd.notna(subj_val) else ''
            #     break

            # 변경: 컬럼 전체 데이터를 분석하여 DB Spec Type 형식으로 추론
            found_type = infer_db_spec_type(col_series)

            # ── INTEGER인 경우 자리수(n) 미리 계산하여 DS_INT_DIGITS에 보관 ─────────
            # DS_TYPE은 INTEGER 그대로 유지
            # save_data_structure_to_template에서 DB Spec이 DECIMAL인 경우에만
            # DECIMAL(n,0)으로 변환하여 엑셀에 기입 + 비교 수행
            int_digits = 0
            if found_type == 'INTEGER':
                for v in col_series.dropna():
                    v_str = str(v).strip().split('.')[0]
                    int_digits = max(int_digits, len(v_str.lstrip('-').lstrip('0') or '0'))
            # ──────────────────────────────────────────────────────────────────────

            # SUBJID는 기존과 동일하게 값이 있는 첫 번째 대상자 기록
            for idx in df.index:
                val = col_series.iloc[idx]
                if pd.isna(val) or str(val).strip() == '' or str(val).strip().lower() == 'nan':
                    continue
                if subjid_col_raw is not None:
                    subj_val = df[subjid_col_raw].iloc[idx]
                    found_subjid = str(subj_val).strip() if pd.notna(subj_val) else ''
                break

            records.append({
                'DOMAIN'       : domain,
                'ITEM ID'      : item_id,
                'DS_TYPE'      : found_type,
                'DS_SUBJID'    : found_subjid,
                'DS_INT_DIGITS': int_digits,  # INTEGER일 때 정수 최대 자리수, 그 외 0
            })

    return pd.DataFrame(records)


def parse_type_spec(type_str: str):
    """
    DB Spec Type 문자열을 파싱하여 (base_type, params) 튜플로 반환합니다.

    예시:
        'DECIMAL(3,1)'   → ('DECIMAL', (3, 1))
        'NVARCHAR(1024)' → ('NVARCHAR', (1024,))
        'NCHAR(5)'       → ('NCHAR', (5,))
        'INTEGER'        → ('INTEGER', ())
        ''               → ('', ())
    """
    type_str = type_str.strip().upper()
    if not type_str:
        return ('', ())

    if '(' in type_str:
        base  = type_str[:type_str.index('(')].strip()
        inner = type_str[type_str.index('(')+1 : type_str.rindex(')')].strip()
        try:
            params = tuple(int(p.strip()) for p in inner.split(','))
        except ValueError:
            params = ()
    else:
        base   = type_str
        params = ()

    return (base, params)


def compare_ds_type(doc_type: str, ds_type: str) -> str:
    """
    DB Spec의 Type과 Dataset에서 추론한 Type을 비교하여 판정 문자열을 반환합니다.

    반환값:
    - 'TRUE'    : Type 일치 + 자리수 완전 동일
    - '확인 필요': Type 일치 + 자리수가 기준 미만 (더 작음)
    - 'FALSE'   : 데이터 없음 / base type 불일치 / 자리수 초과

    Type별 판정 기준:
    - INTEGER:  base type 일치                          → TRUE (자리수 개념 없음)
    - NVARCHAR: 길이 ==                                 → TRUE
                길이 <                                  → 확인 필요
                길이 >                                  → FALSE
    - DECIMAL:  전체자리 == AND 소수자리 ==              → TRUE
                전체자리 ≤ AND 소수자리 ≤ (하나라도 미만) → 확인 필요
                전체자리 > OR 소수자리 >                 → FALSE
    - NCHAR:    Dataset은 항상 NVARCHAR로 추론되므로
                base type 불일치 → FALSE (사람이 길이 검토)

    [수정 이력]
    - 반환 타입 bool → str 변경 ('TRUE' / '확인 필요' / 'FALSE')
    - NVARCHAR, DECIMAL에 자리수 미만 시 '확인 필요' 분기 추가
    """
    if not ds_type:
        return 'FALSE'

    doc_base, doc_params = parse_type_spec(doc_type)
    ds_base,  ds_params  = parse_type_spec(ds_type)

    if not doc_base:
        return 'FALSE'

    if doc_base != ds_base:
        # NCHAR vs NVARCHAR 불일치도 여기서 FALSE 처리됨
        return 'FALSE'

    if doc_base == 'INTEGER':
        return 'TRUE'

    if doc_base == 'NVARCHAR':
        if not doc_params or not ds_params:
            return 'FALSE'
        if ds_params[0] > doc_params[0]:
            return 'FALSE'
        if ds_params[0] == doc_params[0]:
            return 'TRUE'
        return '확인 필요'  # ds 길이 < doc 길이

    if doc_base == 'NCHAR':
        # Dataset 추론이 NVARCHAR로 통일되어 실질적으로 이 분기에 도달하지 않음
        if not doc_params or not ds_params:
            return 'FALSE'
        return 'TRUE' if doc_params[0] == ds_params[0] else 'FALSE'

    if doc_base == 'DECIMAL':
        if not doc_params or not ds_params:
            return 'FALSE'
        if len(doc_params) < 2 or len(ds_params) < 2:
            return 'FALSE'
        doc_total, doc_dec = doc_params[0], doc_params[1]
        ds_total,  ds_dec  = ds_params[0],  ds_params[1]
        if ds_total > doc_total or ds_dec > doc_dec:
            return 'FALSE'
        if ds_total == doc_total and ds_dec == doc_dec:
            return 'TRUE'
        return '확인 필요'  # 이내이지만 하나 이상 미만

    return 'FALSE'


def save_data_structure_to_template(wb, df_doc_full: pd.DataFrame, df_dataset_long: pd.DataFrame):
    """
    템플릿 워크북의 'Data Structure Validation' 시트에
    DB Spec(전체, 필터 없음)과 CDMS Dataset Long format을 비교하여 기입합니다.

    템플릿 구조 (확인된 실제 구조):
        행3: 'Database Specifications'(A~D 병합) | 'Dataset'(E~G 병합) | '확인 결과'(H) | 'Comment'(I)
        행4: Domain | Item ID | Item Label | Type | Domain | Item ID | Type | (병합) | (병합)
        행5~: 데이터 입력 시작

    추가 열 (코드에서 동적 삽입):
        J열: SUBJID (참조 대상자) — 템플릿에는 없지만 J열에 동적으로 추가

    색상 규칙:
        - 확인 결과 FALSE인 경우(데이터 없음 or Type 불일치/초과) → 연분홍(FFD7E9) 하이라이트
        - 확인 결과 TRUE인 경우 → 흰색(기본)

    확인 결과 판정 (compare_ds_type 함수 위임):
        - DS_TYPE 없음(데이터 없음)                           → FALSE
        - DB Spec Type과 base type 불일치                    → FALSE
        - NVARCHAR: ds 길이 ≤ doc 길이                       → TRUE
        - NCHAR:    ds 길이 == doc 길이                      → TRUE (고정 길이)
        - DECIMAL:  ds 전체자리 ≤ doc, ds 소수자리 ≤ doc     → TRUE
        - INTEGER:  base type 일치                           → TRUE
    """
    sheet_name = 'Data Structure Validation'
    if sheet_name not in wb.sheetnames:
        return wb

    ws = wb[sheet_name]

    # ── 스타일 정의 ──────────────────────────────────────────
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    align_center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left     = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # 연분홍: 아무 대상자도 값이 없는 경우
    light_pink_fill = PatternFill(start_color="FFD7E9", end_color="FFD7E9", fill_type="solid")
    # 흰색: 기본 배경
    white_fill        = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    # 연노란색: 확인 필요
    light_yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # ── 열 위치 상수 (템플릿 고정 구조 기반) ─────────────────
    # A=1, B=2, C=3, D=4 → DB Spec 영역 (Domain, Item ID, Item Label, Type)
    # E=5, F=6, G=7      → Dataset 영역  (Domain, Item ID, Type)
    # H=8                → 확인 결과
    # I=9                → Comment
    # J=10               → SUBJID (동적 추가)
    COL_DOC_DOMAIN    = 1   # A: DB Spec - Domain
    COL_DOC_ITEM_ID   = 2   # B: DB Spec - Item ID
    COL_DOC_ITEM_LABEL= 3   # C: DB Spec - Item Label
    COL_DOC_TYPE      = 4   # D: DB Spec - Type
    COL_DS_DOMAIN     = 5   # E: Dataset - Domain
    COL_DS_ITEM_ID    = 6   # F: Dataset - Item ID
    COL_DS_TYPE       = 7   # G: Dataset - Type
    COL_RESULT        = 8   # H: 확인 결과
    COL_COMMENT       = 9   # I: Comment
    COL_SUBJID        = 10  # J: 참조 대상자 (동적 추가)

    # ── J열 헤더 추가 ─────────────────────────────────────────
    # 3행: 병합 없이 단순 레이블
    hdr3 = ws.cell(row=3, column=COL_SUBJID)
    hdr3.value     = 'SUBJID'
    hdr3.border    = thin_border
    hdr3.alignment = align_center

    # 4행: 세부 레이블
    hdr4 = ws.cell(row=4, column=COL_SUBJID)
    hdr4.value     = '참조 대상자'
    hdr4.border    = thin_border
    hdr4.alignment = align_center

    # ── Dataset Long format을 (DOMAIN, ITEM ID) 복합키로 dict화 ──
    # key: (DOMAIN, ITEM_ID)  value: {'DS_TYPE': ..., 'DS_SUBJID': ...}
    ds_lookup = {}
    for _, r in df_dataset_long.iterrows():
        key = (str(r['DOMAIN']).strip().upper(), str(r['ITEM ID']).strip().upper())
        ds_lookup[key] = {
            'DS_TYPE'      : str(r['DS_TYPE']).strip(),
            'DS_SUBJID'    : str(r['DS_SUBJID']).strip(),
            'DS_INT_DIGITS': int(r['DS_INT_DIGITS']) if 'DS_INT_DIGITS' in r.index else 0,
        }

    # ── DB Spec 기준으로 행 기입 (행 수 = DB Spec 행 수와 동일) ──
    START_ROW = 5  # 데이터 시작 행

    for i, doc_row in df_doc_full.reset_index(drop=True).iterrows():
        r = START_ROW + i

        doc_domain     = str(doc_row.get('DOMAIN',     '')).strip()
        doc_item_id    = str(doc_row.get('ITEM ID',    '')).strip()
        doc_item_label = str(doc_row.get('ITEM LABEL', '')).strip()
        doc_type       = str(doc_row.get('TYPE',       '')).strip()

        # Dataset 매칭 조회
        lookup_key = (doc_domain.upper(), doc_item_id.upper())
        ds_info    = ds_lookup.get(lookup_key, None)

        ds_domain  = doc_domain  if ds_info else ''
        ds_item_id = doc_item_id if ds_info else ''
        ds_type    = ds_info['DS_TYPE']       if ds_info else ''
        ds_subjid  = ds_info['DS_SUBJID']     if ds_info else ''
        ds_int_dig = ds_info['DS_INT_DIGITS'] if ds_info else 0

        # ── DECIMAL(doc) vs INTEGER(ds) → DECIMAL(정수자리+doc_dec, doc_dec)으로 변환 ──
        # DB Spec이 DECIMAL이고 Dataset이 INTEGER(소수점 없는 값)인 경우:
        # 소수점 자리수는 DB Spec의 소수자리(doc_dec)를 따르고,
        # 전체 자리수 = 정수부 자리수(ds_int_dig) + doc_dec
        # 예: DB Spec DECIMAL(4,1), Dataset INTEGER(값:180=3자리) → DECIMAL(3+1,1) = DECIMAL(4,1)
        doc_base_check = doc_type.strip().upper().split('(')[0]
        if doc_base_check == 'DECIMAL' and ds_type == 'INTEGER' and ds_int_dig > 0:
            _, doc_params_check = parse_type_spec(doc_type)
            doc_dec = doc_params_check[1] if len(doc_params_check) >= 2 else 0
            ds_type = f'DECIMAL({ds_int_dig + doc_dec},{doc_dec})'

        # 확인 결과 판정: compare_ds_type으로 DB Spec Type과 Dataset Type 비교
        verdict = compare_ds_type(doc_type, ds_type)  # 'TRUE' / '확인 필요' / 'FALSE'

        # 적용할 배경색 결정
        if verdict == 'TRUE':
            fill = white_fill
        elif verdict == '확인 필요':
            fill = light_yellow_fill
        else:
            fill = light_pink_fill

        # ── 셀 기입 헬퍼 ──────────────────────────────────────
        def write_cell(col, value, align=align_center, apply_fill=False):
            cell           = ws.cell(row=r, column=col)
            cell.value     = value if value != '' else None
            cell.border    = thin_border
            cell.alignment = align
            if apply_fill:
                cell.fill = fill

        # A~D: DB Spec 영역 (배경색 없음 — 기준 문서이므로)
        write_cell(COL_DOC_DOMAIN,     doc_domain)
        write_cell(COL_DOC_ITEM_ID,    doc_item_id)
        write_cell(COL_DOC_ITEM_LABEL, doc_item_label, align=align_left)
        write_cell(COL_DOC_TYPE,       doc_type)

        # E~G: Dataset 영역 (TRUE 이외 연분홍)
        write_cell(COL_DS_DOMAIN,  ds_domain,  apply_fill=True)
        write_cell(COL_DS_ITEM_ID, ds_item_id, apply_fill=True)
        write_cell(COL_DS_TYPE,    ds_type,    apply_fill=True)

        # H: 확인 결과 — 'TRUE' / '확인 필요' / 'FALSE'
        result_cell           = ws.cell(row=r, column=COL_RESULT)
        result_cell.value     = verdict
        result_cell.border    = thin_border
        result_cell.alignment = align_center
        if verdict == '확인 필요':
            result_cell.fill = light_yellow_fill
        elif verdict == 'FALSE':
            result_cell.fill = light_pink_fill

        # I: Comment — 빈칸 (human validation)
        comment_cell           = ws.cell(row=r, column=COL_COMMENT)
        comment_cell.value     = None
        comment_cell.border    = thin_border
        comment_cell.alignment = align_center

        # J: 참조 대상자 SUBJID (no_data이면 연분홍)
        write_cell(COL_SUBJID, ds_subjid, apply_fill=True)

    return wb


# ============================================================
# 4. Entry Screen Validation 저장 함수 (기존 유지)
# ============================================================

def save_to_template(template_path, df_doc, df_edc, ver_info,
                     df_doc_full=None, df_dataset_long=None):
    """
    템플릿에 두 가지 시트 결과를 모두 저장합니다.
      - Entry Screen Validation  : 기존 로직 (df_doc / df_edc 사용)
      - Data Structure Validation: 신규 로직 (df_doc_full / df_dataset_long 사용)

    df_doc_full / df_dataset_long 이 None이면 Data Structure 시트는 건너뜁니다.
    """
    if not os.path.exists(template_path):
        return None

    wb = load_workbook(template_path)

    # ── 버전 정보 기입 ────────────────────────────────────────
    # Entry Screen Validation 시트: A2(Blank), A3(DB Spec), A4(Annotated)
    # Data Structure Validation 시트: A2(DB Spec)
    # 형식 예시: "Blank eCRF Version: V1.1" → "V" + 입력값으로 치환
    def write_version(ws, row, col, label_prefix, ver_value):
        """기존 셀 텍스트에서 버전 부분만 교체하여 기입"""
        cell = ws.cell(row=row, column=col)
        ver_str = f"V{ver_value}" if not str(ver_value).upper().startswith('V') else str(ver_value)
        cell.value = f"{label_prefix}{ver_str}"

    entry_ws = wb['Entry Screen Validation'] if 'Entry Screen Validation' in wb.sheetnames else None
    ds_ws    = wb['Data Structure Validation'] if 'Data Structure Validation' in wb.sheetnames else None

    if entry_ws:
        write_version(entry_ws, row=2, col=1,
                      label_prefix="Blank eCRF Version: ",
                      ver_value=ver_info.get('blank', ''))
        write_version(entry_ws, row=3, col=1,
                      label_prefix="Database Specifications Version: ",
                      ver_value=ver_info.get('db', ''))
        write_version(entry_ws, row=4, col=1,
                      label_prefix="Annotated CRF Version: ",
                      ver_value=ver_info.get('annotated', ''))

    if ds_ws:
        write_version(ds_ws, row=2, col=1,
                      label_prefix="Database Specifications Version: ",
                      ver_value=ver_info.get('db', ''))
    # ─────────────────────────────────────────────────────────

    # ── Entry Screen Validation ───────────────────────────────
    target_sheet = 'Entry Screen Validation'
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]

        template_header_row = 6
        doc_col_map = {}
        edc_col_map = {}

        for col_idx in range(1, 31):
            col_name = ws.cell(row=template_header_row, column=col_idx).value
            if col_name:
                col_name = str(col_name).strip().upper()
                if col_idx <= 15:
                    doc_col_map[col_name] = col_idx
                else:
                    edc_col_map[col_name] = col_idx

        res_col_idx = 31
        for col_idx in range(31, ws.max_column + 1):
            if ("확인 결과" in str(ws.cell(row=5, column=col_idx).value or "") or
                    "확인 결과" in str(ws.cell(row=6, column=col_idx).value or "")):
                res_col_idx = col_idx
                break

        red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        df_doc['ORIGINAL_ORDER'] = range(len(df_doc))
        merged = pd.merge(df_doc, df_edc, on='JOIN_KEY', how='outer',
                          suffixes=('_Doc', '_EDC'), indicator=True)
        merged = (merged.sort_values(by=['ORIGINAL_ORDER'], na_position='last')
                        .drop(columns=['ORIGINAL_ORDER']))

        start_row = 7
        for i, row in merged.reset_index(drop=True).iterrows():
            curr_r = start_row + i
            status = row['_merge']
            cols_to_fill = list(doc_col_map.keys())

            mismatches = []
            if status == 'both':
                for cname in cols_to_fill:
                    d_val = str(row.get(f"{cname}_Doc", "")).strip()
                    e_val = str(row.get(f"{cname}_EDC", "")).strip()
                    if d_val != e_val:
                        mismatches.append(cname)

            for cname, col_idx in doc_col_map.items():
                cell           = ws.cell(row=curr_r, column=col_idx)
                cell.value     = row.get(f"{cname}_Doc", "") if status != 'right_only' else ""
                cell.border    = thin_border
                cell.alignment = align_center
                if status == 'left_only' or (status == 'both' and cname in mismatches):
                    cell.fill = red_fill

            for cname, col_idx in edc_col_map.items():
                cell           = ws.cell(row=curr_r, column=col_idx)
                cell.value     = row.get(f"{cname}_EDC", "") if status != 'left_only' else ""
                cell.border    = thin_border
                cell.alignment = align_center
                if status == 'right_only' or (status == 'both' and cname in mismatches):
                    cell.fill = red_fill

            res_text               = "True" if (status == 'both' and not mismatches) else "False"
            cell_res               = ws.cell(row=curr_r, column=res_col_idx)
            cell_res.value         = res_text
            cell_res.border        = thin_border
            cell_res.alignment     = align_center

    # ── Data Structure Validation ─────────────────────────────
    if df_doc_full is not None and df_dataset_long is not None:
        wb = save_data_structure_to_template(wb, df_doc_full, df_dataset_long)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# 5. UI 구성
# ============================================================

col1, col2 = st.columns([4, 15], vertical_alignment="center")
logo_path = "JNPMEDI_original.jpg"

with col1:
    st.image(logo_path, width=200)

with col2:
    st.title("EDC Validation")

st.info("실시간 프리뷰를 통해 컬럼이 올바르게 인식되는지 확인 후 검증을 시작하세요.")

# ── 파일 업로더 3개 ───────────────────────────────────────────
# DB spec 문서, CDMS DB spec, CDMS dataset
col_u1, col_u2, col_u3 = st.columns(3)
with col_u1:
    doc_file_up = st.file_uploader("📂 기준 문서 (DB Spec)", type=['xlsx', 'xls'], key="doc")
with col_u2:
    edc_file_up = st.file_uploader("📂 Entry Screen 비교 대상 (CDMS Export)",
                                   type=['xlsx', 'xls'], key="edc")
with col_u3:
    dataset_file_up = st.file_uploader("📂 Data Structure 비교 대상 (CDMS Dataset)",
                                       type=['xlsx', 'xls'], key="dataset")

# ── 최소 조건: DB Spec + Entry Screen Export ──────────────────
if doc_file_up and edc_file_up:
    st.markdown("---")

    try:
        doc_excel = load_excel_file(doc_file_up)
        edc_excel = load_excel_file(edc_file_up)
    except Exception as e:
        st.error(f"엑셀 파일 로드 중 오류: {e}")
        st.stop()

    c1, c2 = st.columns(2)

    # DB Spec 설정
    with c1:
        st.subheader("📄 DB Spec 설정")
        doc_sheet  = st.selectbox("시트 선택", doc_excel.sheet_names, key="s1")
        doc_header = st.number_input("헤더 행 (Row Index)", min_value=0, value=1, step=1, key="h1")

        doc_df = get_dynamic_preview(doc_excel, doc_sheet, doc_header)
        st.caption(f"▼ '{doc_sheet}' 시트의 {doc_header}번 행을 헤더로 인식한 결과:")
        st.dataframe(doc_df.head(3), use_container_width=True, hide_index=True)

        is_ok, msg, _ = check_columns_status(doc_df)
        st.markdown(
            f'<div class="{"success-box" if is_ok else "error-box"}">{msg}</div>',
            unsafe_allow_html=True
        )
        doc_ready = is_ok

    # Entry Screen Export 설정
    with c2:
        st.subheader("📄 EDC Export 설정 (Entry Screen)")
        edc_sheet  = st.selectbox("시트 선택", edc_excel.sheet_names, key="s2")
        edc_header = st.number_input("헤더 행 (Row Index)", min_value=0, value=0, step=1, key="h2")

        edc_df = get_dynamic_preview(edc_excel, edc_sheet, edc_header)
        st.caption(f"▼ '{edc_sheet}' 시트의 {edc_header}번 행을 헤더로 인식한 결과:")
        st.dataframe(edc_df.head(3), use_container_width=True, hide_index=True)

        is_ok, msg, _ = check_columns_status(edc_df)
        st.markdown(
            f'<div class="{"success-box" if is_ok else "error-box"}">{msg}</div>',
            unsafe_allow_html=True
        )
        edc_ready = is_ok

    # Dataset 파일 상태 표시
    dataset_ready = False
    if dataset_file_up:
        st.markdown("---")
        st.subheader("📄 CDMS Dataset 확인")
        try:
            dataset_excel  = load_excel_file(dataset_file_up)
            domain_sheets  = [s for s in dataset_excel.sheet_names
                              if s.upper() != 'SUBJECT_INFO']
            st.markdown(
                f'<div class="success-box">✅ Dataset 로드 성공 — '
                f'도메인 시트 {len(domain_sheets)}개 인식: '
                f'{", ".join(domain_sheets)}</div>',
                unsafe_allow_html=True
            )
            dataset_ready = True
        except Exception as e:
            st.markdown(
                f'<div class="error-box">⚠️ Dataset 파일 로드 실패: {e}</div>',
                unsafe_allow_html=True
            )
    else:
        st.info("ℹ️ CDMS Dataset 파일을 업로드하면 **Data Structure Validation**도 함께 수행됩니다.")

    st.markdown("---")

    # 버전 정보
    with st.expander("📌 버전 정보 (Optional)", expanded=False):
        v1, v2, v3 = st.columns(3)
        bv = v1.text_input("Blank Ver.", "1.0")
        dv = v2.text_input("DB Spec Ver.", "1.0")
        av = v3.text_input("Annotated Ver.", "1.0")

    if not os.path.exists(TEMPLATE_PATH):
        st.error(f"🚨 중요: 실행 경로에 '{TEMPLATE_PATH}' 파일이 없습니다.")
        btn_disabled = True
    else:
        btn_disabled = not (doc_ready and edc_ready)

    if st.button("🚀 검증 시작 (Start Validation)", type="primary", disabled=btn_disabled):
        with st.status("검증 실행 중 — 잠시 기다려 주세요.", expanded=True) as status:

            # ── DB Spec 로드 ──────────────────────────────────
            df_doc_full = process_data_final(doc_excel, doc_sheet, doc_header)  # 전체 (필터 없음)

            if df_doc_full.empty:
                status.update(label="❌ DB Spec 로드 실패", state="error")
                st.error("DB Spec 데이터를 불러올 수 없습니다.")
                st.stop()

            st.write("📖 DB Spec 로드 - 완료")

            # ── Entry Screen: SYS_ 필터 적용 ─────────────────
            df_doc_entry, df_excluded = apply_sys_layout_filter(df_doc_full.copy(), SYS_LAYOUT_WHITELIST)
            st.write("🔍 Entry Screen SYS_ 필터 적용 - 완료")

            if not df_excluded.empty:
                st.info(
                    f"ℹ️ SYS_ 레이아웃으로 인해 Entry Screen 비교에서 제외된 항목: "
                    f"**{len(df_excluded)}건** (Whitelist 항목은 포함 유지)"
                )
                with st.expander("제외된 항목 확인 (SYS_ 필터)"):
                    st.dataframe(
                        df_excluded[['DOMAIN', 'PAGE', 'VISIT', 'ITEM ID', 'LAYOUT']],
                        use_container_width=True, hide_index=True
                    )

            # ── Entry Screen: EDC Export 로드 ─────────────────
            df_final_edc = process_data_final(edc_excel, edc_sheet, edc_header)

            if df_final_edc.empty:
                status.update(label="❌ EDC Export 로드 실패", state="error")
                st.error("EDC Export 데이터를 불러올 수 없습니다.")
                st.stop()

            # ── EDC Export에도 동일한 SYS_ 필터 적용 ──────────
            df_final_edc, df_edc_excluded = apply_sys_layout_filter(
                df_final_edc, SYS_LAYOUT_WHITELIST
            )
            st.write("📖 EDC Export 로드 및 SYS_ 필터 적용 - 완료")

            if not df_edc_excluded.empty:
                st.info(
                    f"ℹ️ EDC Export에서도 SYS_ 레이아웃으로 제외된 항목: "
                    f"**{len(df_edc_excluded)}건**"
                )

            # ── Data Structure: Dataset Long format 변환 ──────
            df_dataset_long = None
            if dataset_ready:
                df_dataset_long = build_dataset_long(dataset_excel)
                st.write(
                    f"🔄 CDMS Dataset 변환 - 완료 "
                    f"(총 **{len(df_dataset_long)}개** Domain-Item ID 조합 추출)"
                )

            # ── 템플릿에 저장 ─────────────────────────────────
            st.write("📝 템플릿 결과 기입 - 완료")
            ver_info    = {'blank': bv, 'db': dv, 'annotated': av}
            result_file = save_to_template(
                TEMPLATE_PATH,
                df_doc_entry,       # Entry Screen용 (SYS_ 필터 적용)
                df_final_edc,
                ver_info,
                df_doc_full=df_doc_full,            # Data Structure용 (필터 없음)
                df_dataset_long=df_dataset_long,    # None이면 해당 시트 건너뜀
            )

            if result_file:
                status.update(label="🎉 완료!", state="complete")

                summary_parts = ["✅ **Entry Screen Validation** 완료"]
                if df_dataset_long is not None:
                    false_cnt = (df_dataset_long['DS_TYPE'] == '').sum() if not df_dataset_long.empty else 0
                    summary_parts.append(
                        f"✅ **Data Structure Validation** 완료 "
                        f"(데이터 없는 항목: {false_cnt}건 → 연분홍 표시 + FALSE / "
                        f"Type 불일치 항목도 연분홍 표시 + FALSE)"
                    )
                else:
                    summary_parts.append("⚠️ CDMS Dataset 미업로드 → Data Structure Validation 건너뜀")

                st.success("\n\n".join(summary_parts))

                today_str = pd.Timestamp.now().strftime('%Y%m%d')
                file_name = f"EDC Validation List_{today_str}.xlsx"
                st.download_button(
                    label="📥 결과 리포트 다운로드",
                    data=result_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                status.update(label="❌ 템플릿 저장 실패", state="error")
                st.error("결과 파일 생성 중 오류가 발생했습니다.")

else:
    st.info("👆 먼저 상단에서 기준 문서(DB Spec)와 CDMS Export 파일을 업로드해주세요.")