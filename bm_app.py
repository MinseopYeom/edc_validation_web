import streamlit as st
import pandas as pd
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell

# ============================================================
# 1. 페이지 설정
# ============================================================
icon_path = "blue-white.png"
st.set_page_config(page_title="JNPMEDI EDC Validation", page_icon=icon_path, layout="wide")

TEMPLATE_PATH = 'EDC Validation_template.xlsx'

# ============================================================
# [유지보수 포인트] SYS_ 레이아웃 제외 시 포함 예외 목록 (ITEM ID 기준)
# 추후 비교에 포함시켜야 할 ITEM ID가 생기면 이 리스트에 추가하세요.
# ============================================================
SYS_LAYOUT_WHITELIST = [
    "SUBJID",
    # "SITEID",  # 예시: 추후 추가할 경우 이런 식으로 등록
]

st.markdown("""
    <style>
    .stApp { background-color: #F4F7F6; color: #333333; }
    h1, h2, h3, h4, h5, h6, p, span, div, label {
        color: #2c3e50 !important;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div {
        background-color: #ffffff !important; color: #333333 !important;
        border: 1px solid #dcdcdc; border-radius: 8px;
    }
    ul[data-testid="stSelectboxVirtualDropdown"] li {
        color: #333333 !important; background-color: #ffffff !important;
    }
    .stFileUploader, div[data-testid="stExpander"], div[data-testid="stVerticalBlock"] > div {
        background-color: #ffffff; color: #333333 !important;
        border-radius: 10px; padding: 5px;
    }
    .stFileUploader label { font-weight: bold; font-size: 1.1em; }
    .stButton > button, .stDownloadButton > button {
        width: 100%; background-color: #008fd4; color: #ffffff !important;
        font-weight: bold; border: none; padding: 0.6rem; border-radius: 8px;
        transition: all 0.3s ease; box-shadow: 0 2px 4px rgba(0,143,212,0.3);
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background-color: #006fa3; color: #ffffff !important;
        box-shadow: 0 4px 8px rgba(0,111,163,0.4); transform: translateY(-1px);
    }
    .stButton > button:active { transform: translateY(0px); }
    .success-box {
        padding: 15px; background-color: #e3f2fd; color: #0d47a1 !important;
        border-left: 5px solid #008fd4; border-radius: 4px;
        margin-bottom: 15px; font-weight: 600;
    }
    .error-box {
        padding: 15px; background-color: #ffebee; color: #b71c1c !important;
        border-left: 5px solid #d32f2f; border-radius: 4px;
        margin-bottom: 15px; font-weight: 600;
    }
    </style>
""", unsafe_allow_html=True)


# ============================================================
# 2. 공통 유틸 함수
# ============================================================

@st.cache_resource
def load_excel_file(file):
    """파일을 메모리에 로드 (속도 향상)"""
    return pd.ExcelFile(file)


def get_dynamic_preview(excel_file, sheet_name, header_row):
    """사용자가 선택한 행을 헤더로 적용하여 10개 미리보기 생성"""
    try:
        return pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row, nrows=10, dtype=str)
    except Exception:
        return pd.DataFrame()


def check_columns_status(df):
    """필수 컬럼이 식별되는지 진단"""
    if df.empty:
        return False, "데이터 없음", []

    current_cols = [str(c).upper().strip() for c in df.columns]

    rename_map = {
        'VAR NAME': 'ITEM ID', 'VARIABLE NAME': 'ITEM ID', 'VARIABLE': 'ITEM ID',
        'OID': 'ITEM ID', 'ITEMOID': 'ITEM ID', 'QUESTION OID': 'ITEM ID',
        'FORM': 'PAGE', 'FORM OID': 'PAGE', 'FORM NAME': 'PAGE', 'CRF PAGE': 'PAGE',
        'FOLDER': 'VISIT', 'FOLDER OID': 'VISIT', 'EVENT': 'VISIT', 'VISIT NAME': 'VISIT',
        'DATASET': 'DOMAIN', 'LB DOMAIN': 'DOMAIN', 'DOMAIN NAME': 'DOMAIN',
        'VER.': 'VERSION', 'VER': 'VERSION', 'CRF_VERSION': 'VERSION', 'CRF VERSION': 'VERSION',
    }

    mapped_cols = set()
    for col in current_cols:
        if col in rename_map:
            mapped_cols.add(rename_map[col])
        elif col in {'DOMAIN', 'PAGE', 'VISIT', 'ITEM ID'}:
            mapped_cols.add(col)

    required = {'DOMAIN', 'PAGE', 'VISIT', 'ITEM ID'}
    missing = required - mapped_cols

    if not missing:
        return True, "✅ 필수 컬럼 자동 인식 성공!", []
    else:
        return False, f"⚠️ 필수 컬럼 미식별: {', '.join(missing)}", list(missing)


def apply_sys_layout_filter(df, whitelist):
    """
    DB Spec에서 SYS_ 레이아웃 행을 필터링합니다.
    - LAYOUT이 'SYS_'로 시작하면 제외
    - 단 ITEM ID가 whitelist에 있으면 포함 유지
    """
    if 'LAYOUT' not in df.columns:
        return df, pd.DataFrame()

    whitelist_upper = [item.upper().strip() for item in whitelist]
    is_sys = df['LAYOUT'].str.upper().str.startswith('SYS_')
    is_whitelisted = df['ITEM ID'].str.upper().isin(whitelist_upper)
    exclude_mask = is_sys & ~is_whitelisted

    return df[~exclude_mask].reset_index(drop=True), df[exclude_mask].reset_index(drop=True)


def process_data_final(excel_file, sheet_name, header_row):
    """DB Spec 파일을 읽어 표준화된 DataFrame으로 반환"""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row, dtype=str)
        df.columns = [str(c).upper().strip() for c in df.columns]

        rename_map = {
            'VAR NAME': 'ITEM ID', 'VARIABLE NAME': 'ITEM ID', 'VARIABLE': 'ITEM ID',
            'OID': 'ITEM ID', 'ITEMOID': 'ITEM ID',
            'FORM': 'PAGE', 'FORM OID': 'PAGE', 'FORM NAME': 'PAGE', 'CRF PAGE': 'PAGE',
            'FOLDER': 'VISIT', 'FOLDER OID': 'VISIT', 'EVENT': 'VISIT',
            'DATASET': 'DOMAIN', 'LB DOMAIN': 'DOMAIN',
            'VER.': 'VERSION', 'VER': 'VERSION', 'CRF_VERSION': 'VERSION', 'CRF VERSION': 'VERSION',
        }
        df = df.rename(columns=rename_map)

        std_cols = ['DOMAIN', 'DOMAIN LABEL', 'PAGE', 'PAGE LABEL', 'VISIT',
                    'ITEM ID', 'ITEM LABEL', 'ITEM SEQ', 'VERSION', 'CODE',
                    'LAYOUT', 'TYPE', 'MAX_LEN', 'MIN_VAL', 'MAX_VAL']

        for col in std_cols:
            if col not in df.columns:
                df[col] = ""
            df[col] = (df[col].fillna("").astype(str)
                       .apply(lambda x: x.replace('.0', '').strip() if x.endswith('.0') else x.strip()))

        df['JOIN_KEY'] = (df['DOMAIN'] + df['PAGE'] + df['VISIT'] + df['ITEM ID']
                          ).str.replace(r'\s+', '', regex=True).str.upper()

        df = df[df['JOIN_KEY'].str.len() > 1]
        df = df.drop_duplicates(subset=['JOIN_KEY'])
        return df
    except Exception:
        return pd.DataFrame()


# ============================================================
# 3. Data Structure Validation 관련 함수
# ============================================================

def parse_item_id(col_name: str) -> str:
    """
    'ITEMID:LABEL' 형태의 컬럼명에서 ITEM ID 부분만 추출합니다.
    ':' 가 없으면 컬럼명 그대로 반환합니다.
    """
    return col_name.split(':')[0].strip().upper()


def dtype_to_type_str(dtype) -> str:
    """
    pandas dtype을 사람이 읽기 쉬운 Type 문자열로 변환합니다.
    DB Spec의 TYPE 컬럼과 비교하기 위한 참고값입니다.
    """
    dtype_str = str(dtype)
    if 'datetime' in dtype_str:
        return 'datetime'
    elif 'int' in dtype_str:
        return 'integer'
    elif 'float' in dtype_str:
        return 'float'
    else:
        return 'text'


def build_dataset_long(dataset_excel: pd.ExcelFile) -> pd.DataFrame:
    """
    CDMS Dataset 엑셀의 모든 도메인 시트를 읽어 Long format DataFrame으로 변환합니다.

    변환 규칙:
    - 시트명 = DOMAIN
    - 컬럼명 'ITEMID:LABEL' → ITEM ID는 ':' 앞 부분만 추출
    - 모든 컬럼을 Item ID로 처리 (제외 없음)
    - 각 Item ID에 대해 값이 실제로 존재하는(non-null) 첫 번째 행의
      실제 셀 값을 Type으로, 해당 행의 SUBJID를 참조 대상자로 기록
    - 모든 대상자에게 값이 없는 경우 DS_TYPE = '', DS_SUBJID = '' 으로 기록

    Returns:
        DataFrame with columns: [DOMAIN, ITEM ID, DS_TYPE, DS_SUBJID]
    """
    skip_sheets = {'SUBJECT_INFO'}
    records = []

    for sheet in dataset_excel.sheet_names:
        if sheet.upper() in skip_sheets:
            continue

        domain = sheet.strip().upper()

        try:
            df = pd.read_excel(dataset_excel, sheet_name=sheet)
        except Exception:
            continue

        if df.empty:
            continue

        # SUBJID 컬럼 원본명 찾기 (SUBJID:xxx 형태일 수 있음)
        subjid_col_raw = None
        for c in df.columns:
            if parse_item_id(c) == 'SUBJID':
                subjid_col_raw = c
                break

        # 모든 컬럼을 Item ID로 처리
        for raw_col in df.columns:
            item_id = parse_item_id(raw_col)

            col_series   = df[raw_col]
            found_subjid = ''
            found_type   = ''

            for idx in df.index:
                val = col_series.iloc[idx]
                if pd.isna(val) or str(val).strip() == '' or str(val).strip().lower() == 'nan':
                    continue
                # 값이 있는 첫 번째 대상자의 실제 셀 값을 그대로 사용
                found_type = str(val).strip()
                if subjid_col_raw is not None:
                    subj_val = df[subjid_col_raw].iloc[idx]
                    found_subjid = str(subj_val).strip() if pd.notna(subj_val) else ''
                break

            records.append({
                'DOMAIN'   : domain,
                'ITEM ID'  : item_id,
                'DS_TYPE'  : found_type,
                'DS_SUBJID': found_subjid,
            })

    return pd.DataFrame(records)


def save_data_structure_to_template(wb, df_doc_full: pd.DataFrame, df_dataset_long: pd.DataFrame):
    """
    템플릿 워크북의 'Data Structure Validation' 시트에
    DB Spec(전체, 필터 없음)과 CDMS Dataset Long format을 비교하여 기입합니다.

    템플릿 구조 (확인된 실제 구조):
        행3: 'Database Specifications'(A~D 병합) | 'Dataset'(E~G 병합) | '확인 결과'(H) | 'Comment'(I)
        행4: Domain | Item ID | Item Label | Type | Domain | Item ID | Type | (병합) | (병합)
        행5~: 데이터 입력 시작

    추가 열 (코드에서 동적 삽입):
        J열: SUBJID (참조 대상자) — 템플릿에는 없지만 J열에 동적으로 추가

    색상 규칙:
        - Dataset에서 해당 값이 아예 없는 경우(DS_TYPE이 빈값) → 연분홍(FFD7E9) 하이라이트
        - 확인 결과: 값이 없는 경우 'FALSE', 있는 경우 빈칸(human validation)
    """
    sheet_name = 'Data Structure Validation'
    if sheet_name not in wb.sheetnames:
        return wb

    ws = wb[sheet_name]

    # ── 스타일 정의 ──────────────────────────────────────────
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    align_center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left     = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # 연분홍: 아무 대상자도 값이 없는 경우
    light_pink_fill = PatternFill(start_color="FFD7E9", end_color="FFD7E9", fill_type="solid")
    # 흰색: 기본 배경
    white_fill      = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ── 열 위치 상수 (템플릿 고정 구조 기반) ─────────────────
    # A=1, B=2, C=3, D=4 → DB Spec 영역 (Domain, Item ID, Item Label, Type)
    # E=5, F=6, G=7      → Dataset 영역  (Domain, Item ID, Type)
    # H=8                → 확인 결과
    # I=9                → Comment
    # J=10               → SUBJID (동적 추가)
    COL_DOC_DOMAIN    = 1   # A: DB Spec - Domain
    COL_DOC_ITEM_ID   = 2   # B: DB Spec - Item ID
    COL_DOC_ITEM_LABEL= 3   # C: DB Spec - Item Label
    COL_DOC_TYPE      = 4   # D: DB Spec - Type
    COL_DS_DOMAIN     = 5   # E: Dataset - Domain
    COL_DS_ITEM_ID    = 6   # F: Dataset - Item ID
    COL_DS_TYPE       = 7   # G: Dataset - Type
    COL_RESULT        = 8   # H: 확인 결과
    COL_COMMENT       = 9   # I: Comment
    COL_SUBJID        = 10  # J: 참조 대상자 (동적 추가)

    # ── J열 헤더 추가 ─────────────────────────────────────────
    # 3행: 병합 없이 단순 레이블
    hdr3 = ws.cell(row=3, column=COL_SUBJID)
    hdr3.value     = 'SUBJID'
    hdr3.border    = thin_border
    hdr3.alignment = align_center

    # 4행: 세부 레이블
    hdr4 = ws.cell(row=4, column=COL_SUBJID)
    hdr4.value     = '참조 대상자'
    hdr4.border    = thin_border
    hdr4.alignment = align_center

    # ── Dataset Long format을 (DOMAIN, ITEM ID) 복합키로 dict화 ──
    # key: (DOMAIN, ITEM_ID)  value: {'DS_TYPE': ..., 'DS_SUBJID': ...}
    ds_lookup = {}
    for _, r in df_dataset_long.iterrows():
        key = (str(r['DOMAIN']).strip().upper(), str(r['ITEM ID']).strip().upper())
        ds_lookup[key] = {
            'DS_TYPE'  : str(r['DS_TYPE']).strip(),
            'DS_SUBJID': str(r['DS_SUBJID']).strip(),
        }

    # ── DB Spec 기준으로 행 기입 (행 수 = DB Spec 행 수와 동일) ──
    START_ROW = 5  # 데이터 시작 행

    for i, doc_row in df_doc_full.reset_index(drop=True).iterrows():
        r = START_ROW + i

        doc_domain     = str(doc_row.get('DOMAIN',     '')).strip()
        doc_item_id    = str(doc_row.get('ITEM ID',    '')).strip()
        doc_item_label = str(doc_row.get('ITEM LABEL', '')).strip()
        doc_type       = str(doc_row.get('TYPE',       '')).strip()

        # Dataset 매칭 조회
        lookup_key = (doc_domain.upper(), doc_item_id.upper())
        ds_info    = ds_lookup.get(lookup_key, None)

        ds_domain  = doc_domain  if ds_info else ''
        ds_item_id = doc_item_id if ds_info else ''
        ds_type    = ds_info['DS_TYPE']   if ds_info else ''
        ds_subjid  = ds_info['DS_SUBJID'] if ds_info else ''

        # 값이 없는 경우(아무 대상자도 해당 item에 데이터 없음) 판별
        no_data = (ds_type == '')

        # 적용할 배경색 결정
        fill = light_pink_fill if no_data else white_fill

        # ── 셀 기입 헬퍼 ──────────────────────────────────────
        def write_cell(col, value, align=align_center, apply_fill=False):
            cell           = ws.cell(row=r, column=col)
            cell.value     = value if value != '' else None
            cell.border    = thin_border
            cell.alignment = align
            if apply_fill:
                cell.fill = fill

        # A~D: DB Spec 영역 (배경색 없음 — 기준 문서이므로)
        write_cell(COL_DOC_DOMAIN,     doc_domain)
        write_cell(COL_DOC_ITEM_ID,    doc_item_id)
        write_cell(COL_DOC_ITEM_LABEL, doc_item_label, align=align_left)
        write_cell(COL_DOC_TYPE,       doc_type)

        # E~G: Dataset 영역 (no_data이면 연분홍)
        write_cell(COL_DS_DOMAIN,  ds_domain,  apply_fill=True)
        write_cell(COL_DS_ITEM_ID, ds_item_id, apply_fill=True)
        write_cell(COL_DS_TYPE,    ds_type,    apply_fill=True)

        # H: 확인 결과 — 값 없으면 FALSE, 있으면 빈칸
        result_cell           = ws.cell(row=r, column=COL_RESULT)
        result_cell.value     = 'FALSE' if no_data else None
        result_cell.border    = thin_border
        result_cell.alignment = align_center
        if no_data:
            result_cell.fill = light_pink_fill

        # I: Comment — 빈칸 (human validation)
        comment_cell           = ws.cell(row=r, column=COL_COMMENT)
        comment_cell.value     = None
        comment_cell.border    = thin_border
        comment_cell.alignment = align_center

        # J: 참조 대상자 SUBJID (no_data이면 연분홍)
        write_cell(COL_SUBJID, ds_subjid, apply_fill=True)

    return wb


# ============================================================
# 4. Entry Screen Validation 저장 함수 (기존 유지)
# ============================================================

def save_to_template(template_path, df_doc, df_edc, ver_info,
                     df_doc_full=None, df_dataset_long=None):
    """
    템플릿에 두 가지 시트 결과를 모두 저장합니다.
      - Entry Screen Validation  : 기존 로직 (df_doc / df_edc 사용)
      - Data Structure Validation: 신규 로직 (df_doc_full / df_dataset_long 사용)

    df_doc_full / df_dataset_long 이 None이면 Data Structure 시트는 건너뜁니다.
    """
    if not os.path.exists(template_path):
        return None

    wb = load_workbook(template_path)

    # ── 버전 정보 기입 ────────────────────────────────────────
    # Entry Screen Validation 시트: A2(Blank), A3(DB Spec), A4(Annotated)
    # Data Structure Validation 시트: A2(DB Spec)
    # 형식 예시: "Blank eCRF Version: V1.1" → "V" + 입력값으로 치환
    def write_version(ws, row, col, label_prefix, ver_value):
        """기존 셀 텍스트에서 버전 부분만 교체하여 기입"""
        cell = ws.cell(row=row, column=col)
        ver_str = f"V{ver_value}" if not str(ver_value).upper().startswith('V') else str(ver_value)
        cell.value = f"{label_prefix}{ver_str}"

    entry_ws = wb['Entry Screen Validation'] if 'Entry Screen Validation' in wb.sheetnames else None
    ds_ws    = wb['Data Structure Validation'] if 'Data Structure Validation' in wb.sheetnames else None

    if entry_ws:
        write_version(entry_ws, row=2, col=1,
                      label_prefix="Blank eCRF Version: ",
                      ver_value=ver_info.get('blank', ''))
        write_version(entry_ws, row=3, col=1,
                      label_prefix="Database Specifications Version: ",
                      ver_value=ver_info.get('db', ''))
        write_version(entry_ws, row=4, col=1,
                      label_prefix="Annotated CRF Version: ",
                      ver_value=ver_info.get('annotated', ''))

    if ds_ws:
        write_version(ds_ws, row=2, col=1,
                      label_prefix="Database Specifications Version: ",
                      ver_value=ver_info.get('db', ''))
    # ─────────────────────────────────────────────────────────

    # ── Entry Screen Validation ───────────────────────────────
    target_sheet = 'Entry Screen Validation'
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]

        template_header_row = 6
        doc_col_map = {}
        edc_col_map = {}

        for col_idx in range(1, 31):
            col_name = ws.cell(row=template_header_row, column=col_idx).value
            if col_name:
                col_name = str(col_name).strip().upper()
                if col_idx <= 15:
                    doc_col_map[col_name] = col_idx
                else:
                    edc_col_map[col_name] = col_idx

        res_col_idx = 31
        for col_idx in range(31, ws.max_column + 1):
            if ("확인 결과" in str(ws.cell(row=5, column=col_idx).value or "") or
                    "확인 결과" in str(ws.cell(row=6, column=col_idx).value or "")):
                res_col_idx = col_idx
                break

        red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        df_doc['ORIGINAL_ORDER'] = range(len(df_doc))
        merged = pd.merge(df_doc, df_edc, on='JOIN_KEY', how='outer',
                          suffixes=('_Doc', '_EDC'), indicator=True)
        merged = (merged.sort_values(by=['ORIGINAL_ORDER'], na_position='last')
                        .drop(columns=['ORIGINAL_ORDER']))

        start_row = 7
        for i, row in merged.reset_index(drop=True).iterrows():
            curr_r = start_row + i
            status = row['_merge']
            cols_to_fill = list(doc_col_map.keys())

            mismatches = []
            if status == 'both':
                for cname in cols_to_fill:
                    d_val = str(row.get(f"{cname}_Doc", "")).strip()
                    e_val = str(row.get(f"{cname}_EDC", "")).strip()
                    if d_val != e_val:
                        mismatches.append(cname)

            for cname, col_idx in doc_col_map.items():
                cell           = ws.cell(row=curr_r, column=col_idx)
                cell.value     = row.get(f"{cname}_Doc", "") if status != 'right_only' else ""
                cell.border    = thin_border
                cell.alignment = align_center
                if status == 'left_only' or (status == 'both' and cname in mismatches):
                    cell.fill = red_fill

            for cname, col_idx in edc_col_map.items():
                cell           = ws.cell(row=curr_r, column=col_idx)
                cell.value     = row.get(f"{cname}_EDC", "") if status != 'left_only' else ""
                cell.border    = thin_border
                cell.alignment = align_center
                if status == 'right_only' or (status == 'both' and cname in mismatches):
                    cell.fill = red_fill

            res_text               = "True" if (status == 'both' and not mismatches) else "False"
            cell_res               = ws.cell(row=curr_r, column=res_col_idx)
            cell_res.value         = res_text
            cell_res.border        = thin_border
            cell_res.alignment     = align_center

    # ── Data Structure Validation ─────────────────────────────
    if df_doc_full is not None and df_dataset_long is not None:
        wb = save_data_structure_to_template(wb, df_doc_full, df_dataset_long)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# 5. UI 구성
# ============================================================

col1, col2 = st.columns([4, 15], vertical_alignment="center")
logo_path = "JNPMEDI_original.jpg"

with col1:
    st.image(logo_path, width=200)

with col2:
    st.title("EDC Validation")

st.info("실시간 프리뷰를 통해 컬럼이 올바르게 인식되는지 확인 후 검증을 시작하세요.")

# ── 파일 업로더 3개 ───────────────────────────────────────────
# DB spec 문서, CDMS DB spec, CDMS dataset
col_u1, col_u2, col_u3 = st.columns(3)
with col_u1:
    doc_file_up = st.file_uploader("📂 기준 문서 (DB Spec)", type=['xlsx', 'xls'], key="doc")
with col_u2:
    edc_file_up = st.file_uploader("📂 Entry Screen 비교 대상 (CDMS Export)",
                                   type=['xlsx', 'xls'], key="edc")
with col_u3:
    dataset_file_up = st.file_uploader("📂 Data Structure 비교 대상 (CDMS Dataset)",
                                       type=['xlsx', 'xls'], key="dataset")

# ── 최소 조건: DB Spec + Entry Screen Export ──────────────────
if doc_file_up and edc_file_up:
    st.markdown("---")

    try:
        doc_excel = load_excel_file(doc_file_up)
        edc_excel = load_excel_file(edc_file_up)
    except Exception as e:
        st.error(f"엑셀 파일 로드 중 오류: {e}")
        st.stop()

    c1, c2 = st.columns(2)

    # DB Spec 설정
    with c1:
        st.subheader("📄 DB Spec 설정")
        doc_sheet  = st.selectbox("시트 선택", doc_excel.sheet_names, key="s1")
        doc_header = st.number_input("헤더 행 (Row Index)", min_value=0, value=1, step=1, key="h1")

        doc_df = get_dynamic_preview(doc_excel, doc_sheet, doc_header)
        st.caption(f"▼ '{doc_sheet}' 시트의 {doc_header}번 행을 헤더로 인식한 결과:")
        st.dataframe(doc_df.head(3), use_container_width=True, hide_index=True)

        is_ok, msg, _ = check_columns_status(doc_df)
        st.markdown(
            f'<div class="{"success-box" if is_ok else "error-box"}">{msg}</div>',
            unsafe_allow_html=True
        )
        doc_ready = is_ok

    # Entry Screen Export 설정
    with c2:
        st.subheader("📄 EDC Export 설정 (Entry Screen)")
        edc_sheet  = st.selectbox("시트 선택", edc_excel.sheet_names, key="s2")
        edc_header = st.number_input("헤더 행 (Row Index)", min_value=0, value=0, step=1, key="h2")

        edc_df = get_dynamic_preview(edc_excel, edc_sheet, edc_header)
        st.caption(f"▼ '{edc_sheet}' 시트의 {edc_header}번 행을 헤더로 인식한 결과:")
        st.dataframe(edc_df.head(3), use_container_width=True, hide_index=True)

        is_ok, msg, _ = check_columns_status(edc_df)
        st.markdown(
            f'<div class="{"success-box" if is_ok else "error-box"}">{msg}</div>',
            unsafe_allow_html=True
        )
        edc_ready = is_ok

    # Dataset 파일 상태 표시
    dataset_ready = False
    if dataset_file_up:
        st.markdown("---")
        st.subheader("📄 CDMS Dataset 확인")
        try:
            dataset_excel  = load_excel_file(dataset_file_up)
            domain_sheets  = [s for s in dataset_excel.sheet_names
                              if s.upper() != 'SUBJECT_INFO']
            st.markdown(
                f'<div class="success-box">✅ Dataset 로드 성공 — '
                f'도메인 시트 {len(domain_sheets)}개 인식: '
                f'{", ".join(domain_sheets)}</div>',
                unsafe_allow_html=True
            )
            dataset_ready = True
        except Exception as e:
            st.markdown(
                f'<div class="error-box">⚠️ Dataset 파일 로드 실패: {e}</div>',
                unsafe_allow_html=True
            )
    else:
        st.info("ℹ️ CDMS Dataset 파일을 업로드하면 **Data Structure Validation**도 함께 수행됩니다.")

    st.markdown("---")

    # 버전 정보
    with st.expander("📌 버전 정보 (Optional)", expanded=False):
        v1, v2, v3 = st.columns(3)
        bv = v1.text_input("Blank Ver.", "1.0")
        dv = v2.text_input("DB Spec Ver.", "1.0")
        av = v3.text_input("Annotated Ver.", "1.0")

    if not os.path.exists(TEMPLATE_PATH):
        st.error(f"🚨 중요: 실행 경로에 '{TEMPLATE_PATH}' 파일이 없습니다.")
        btn_disabled = True
    else:
        btn_disabled = not (doc_ready and edc_ready)

    if st.button("🚀 검증 시작 (Start Validation)", type="primary", disabled=btn_disabled):
        with st.status("검증 실행 중 — 잠시 기다려 주세요.", expanded=True) as status:

            # ── DB Spec 로드 ──────────────────────────────────
            df_doc_full = process_data_final(doc_excel, doc_sheet, doc_header)  # 전체 (필터 없음)

            if df_doc_full.empty:
                status.update(label="❌ DB Spec 로드 실패", state="error")
                st.error("DB Spec 데이터를 불러올 수 없습니다.")
                st.stop()

            st.write("📖 DB Spec 로드 - 완료")

            # ── Entry Screen: SYS_ 필터 적용 ─────────────────
            df_doc_entry, df_excluded = apply_sys_layout_filter(df_doc_full.copy(), SYS_LAYOUT_WHITELIST)
            st.write("🔍 Entry Screen SYS_ 필터 적용 - 완료")

            if not df_excluded.empty:
                st.info(
                    f"ℹ️ SYS_ 레이아웃으로 인해 Entry Screen 비교에서 제외된 항목: "
                    f"**{len(df_excluded)}건** (Whitelist 항목은 포함 유지)"
                )
                with st.expander("제외된 항목 확인 (SYS_ 필터)"):
                    st.dataframe(
                        df_excluded[['DOMAIN', 'PAGE', 'VISIT', 'ITEM ID', 'LAYOUT']],
                        use_container_width=True, hide_index=True
                    )

            # ── Entry Screen: EDC Export 로드 ─────────────────
            df_final_edc = process_data_final(edc_excel, edc_sheet, edc_header)

            if df_final_edc.empty:
                status.update(label="❌ EDC Export 로드 실패", state="error")
                st.error("EDC Export 데이터를 불러올 수 없습니다.")
                st.stop()

            # ── EDC Export에도 동일한 SYS_ 필터 적용 ──────────
            df_final_edc, df_edc_excluded = apply_sys_layout_filter(
                df_final_edc, SYS_LAYOUT_WHITELIST
            )
            st.write("📖 EDC Export 로드 및 SYS_ 필터 적용 - 완료")

            if not df_edc_excluded.empty:
                st.info(
                    f"ℹ️ EDC Export에서도 SYS_ 레이아웃으로 제외된 항목: "
                    f"**{len(df_edc_excluded)}건**"
                )

            # ── Data Structure: Dataset Long format 변환 ──────
            df_dataset_long = None
            if dataset_ready:
                df_dataset_long = build_dataset_long(dataset_excel)
                st.write(
                    f"🔄 CDMS Dataset 변환 - 완료 "
                    f"(총 **{len(df_dataset_long)}개** Domain-Item ID 조합 추출)"
                )

            # ── 템플릿에 저장 ─────────────────────────────────
            st.write("📝 템플릿 결과 기입 - 완료")
            ver_info    = {'blank': bv, 'db': dv, 'annotated': av}
            result_file = save_to_template(
                TEMPLATE_PATH,
                df_doc_entry,       # Entry Screen용 (SYS_ 필터 적용)
                df_final_edc,
                ver_info,
                df_doc_full=df_doc_full,            # Data Structure용 (필터 없음)
                df_dataset_long=df_dataset_long,    # None이면 해당 시트 건너뜀
            )

            if result_file:
                status.update(label="🎉 완료!", state="complete")

                summary_parts = ["✅ **Entry Screen Validation** 완료"]
                if df_dataset_long is not None:
                    no_data_cnt = (df_dataset_long['DS_TYPE'] == '').sum() if not df_dataset_long.empty else 0
                    summary_parts.append(
                        f"✅ **Data Structure Validation** 완료 "
                        f"(데이터 없는 항목: {no_data_cnt}건 → 연분홍 표시 + FALSE)"
                    )
                else:
                    summary_parts.append("⚠️ CDMS Dataset 미업로드 → Data Structure Validation 건너뜀")

                st.success("\n\n".join(summary_parts))

                today_str = pd.Timestamp.now().strftime('%Y%m%d')
                file_name = f"EDC Validation List_{today_str}.xlsx"
                st.download_button(
                    label="📥 결과 리포트 다운로드",
                    data=result_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                status.update(label="❌ 템플릿 저장 실패", state="error")
                st.error("결과 파일 생성 중 오류가 발생했습니다.")

else:
    st.info("👆 먼저 상단에서 기준 문서(DB Spec)와 CDMS Export 파일을 업로드해주세요.")