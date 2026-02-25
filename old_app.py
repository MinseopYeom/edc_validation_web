import streamlit as st
import pandas as pd
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell

# --- 1. 페이지 설정 ---
icon_path = "blue-white.png"
st.set_page_config(page_title="JNPMEDI EDC Validation", page_icon=icon_path, layout="wide")

TEMPLATE_PATH = 'EDC Validation_template.xlsx'

# ============================================================
# [유지보수 포인트] SYS_ 레이아웃 제외 시 포함 예외 목록 (ITEM ID 기준)
# 추후 비교에 포함시켜야 할 ITEM ID가 생기면 이 리스트에 추가.
SYS_LAYOUT_WHITELIST = [
    "SUBJID",
    # "SITEID",   # 예시: 추후 추가할 경우 이런 식으로 등록
]

st.markdown("""
    <style>
    /* 1. 전체 앱 배경 */
    .stApp {
        background-color: #F4F7F6;
        color: #333333;
    }

    /* 2. 텍스트 가독성 강제 해결 (흰색 글씨 방지) */
    h1, h2, h3, h4, h5, h6, p, span, div, label {
        color: #2c3e50 !important;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    
    /* 3. 입력창 디자인 */
    .stTextInput > div > div > input, 
    .stNumberInput > div > div > input ,
    .stSelectbox > div > div {
        background-color: #ffffff !important;
        color: #333333 !important;
        border: 1px solid #dcdcdc;
        border-radius: 8px;
    }
    ul[data-testid="stSelectboxVirtualDropdown"] li {
        color: #333333 !important;
        background-color: #ffffff !important;
    }

    /* 4. 파일 업로더 & 설정 박스 */
    .stFileUploader, div[data-testid="stExpander"], div[data-testid="stVerticalBlock"] > div {
        background-color: #ffffff;
        color: #333333 !important;
        border-radius: 10px;
        padding: 5px;
    }
    .stFileUploader label {
        font-weight: bold;
        font-size: 1.1em;
    }

    /* 5. 버튼 디자인 */
    .stButton > button, .stDownloadButton > button {
        width: 100%;
        background-color: #008fd4;
        color: #ffffff !important;
        font-weight: bold;
        border: none;
        padding: 0.6rem;
        border-radius: 8px;
        transition: all 0.3s ease;
        box-shadow: 0 2px 4px rgba(0,143,212, 0.3);
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background-color: #006fa3;
        color: #ffffff !important;
        box-shadow: 0 4px 8px rgba(0,111,163, 0.4);
        transform: translateY(-1px);
    }
    .stButton > button:active {
        transform: translateY(0px);
    }

    /* 6. 상태 메시지 박스 */
    .success-box {
        padding: 15px;
        background-color: #e3f2fd;
        color: #0d47a1 !important;
        border-left: 5px solid #008fd4;
        border-radius: 4px;
        margin-bottom: 15px;
        font-weight: 600;
    }
    .error-box {
        padding: 15px;
        background-color: #ffebee;
        color: #b71c1c !important;
        border-left: 5px solid #d32f2f;
        border-radius: 4px;
        margin-bottom: 15px;
        font-weight: 600;
    }
    </style>
""", unsafe_allow_html=True)

# --- 2. 핵심 로직 ---

@st.cache_resource
def load_excel_file(file):
    """파일을 메모리에 로드 (속도 향상) - Resource 캐싱 사용"""
    return pd.ExcelFile(file)

def get_dynamic_preview(excel_file, sheet_name, header_row):
    """사용자가 선택한 행을 헤더로 적용하여 미리보기 생성"""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row, nrows=5, dtype=str)
        return df
    except Exception as e:
        return pd.DataFrame()

def check_columns_status(df):
    """필수 컬럼이 식별되는지 진단"""
    if df.empty:
        return False, "데이터 없음", []

    current_cols = [str(c).upper().strip() for c in df.columns]
    
    rename_map = {
        'VAR NAME': 'ITEM ID', 'VARIABLE NAME': 'ITEM ID', 'VARIABLE': 'ITEM ID', 'OID': 'ITEM ID', 
        'ITEMOID': 'ITEM ID', 'QUESTION OID': 'ITEM ID',
        'FORM': 'PAGE', 'FORM OID': 'PAGE', 'FORM NAME': 'PAGE', 'CRF PAGE': 'PAGE',
        'FOLDER': 'VISIT', 'FOLDER OID': 'VISIT', 'EVENT': 'VISIT', 'VISIT NAME': 'VISIT',
        'DATASET': 'DOMAIN', 'LB DOMAIN': 'DOMAIN', 'DOMAIN NAME': 'DOMAIN',
        'VER.': 'VERSION', 
        'VER': 'VERSION', 
        'CRF_VERSION': 'VERSION', 
        'CRF VERSION': 'VERSION', 
    }
    
    mapped_cols = set()
    for col in current_cols:
        if col in rename_map:
            mapped_cols.add(rename_map[col])
        elif col in ['DOMAIN', 'PAGE', 'VISIT', 'ITEM ID']:
            mapped_cols.add(col)
            
    required = {'DOMAIN', 'PAGE', 'VISIT', 'ITEM ID'}
    missing = required - mapped_cols
    
    if not missing:
        return True, "✅ 필수 컬럼 자동 인식 성공!", []
    else:
        return False, f"⚠️ 필수 컬럼 미식별: {', '.join(missing)}", list(missing)


def apply_sys_layout_filter(df, whitelist):
    """
    [신규 함수] DB Spec 데이터에서 SYS_ 레이아웃 행을 필터링합니다.

    규칙:
      - LAYOUT 컬럼 값이 "SYS_"로 시작하는 행은 비교 대상에서 제외합니다.
      - 단, ITEM ID 값이 whitelist(SYS_LAYOUT_WHITELIST)에 포함된 경우는
        SYS_ 레이아웃이더라도 비교 대상에 포함합니다.

    Args:
        df       : process_data_final()을 거친 표준화된 DataFrame
        whitelist: 예외적으로 포함시킬 ITEM ID 목록 (SYS_LAYOUT_WHITELIST)

    Returns:
        filtered_df : 필터링 적용된 DataFrame
        excluded_df : 제외된 행들의 DataFrame (로그/확인용)
    """
    # LAYOUT 컬럼이 아예 없는 경우 필터링 없이 원본 반환
    if 'LAYOUT' not in df.columns:
        return df, pd.DataFrame()

    # whitelist 대소문자 통일 (ITEM ID는 이미 strip 처리되어 있음)
    whitelist_upper = [item.upper().strip() for item in whitelist]

    # SYS_ 시작 여부 판별
    is_sys = df['LAYOUT'].str.upper().str.startswith('SYS_')

    # Whitelist에 있는 ITEM ID 여부 판별
    is_whitelisted = df['ITEM ID'].str.upper().isin(whitelist_upper)

    # 제외 조건: SYS_로 시작하면서 whitelist에 없는 경우
    exclude_mask = is_sys & ~is_whitelisted

    filtered_df  = df[~exclude_mask].reset_index(drop=True)
    excluded_df  = df[exclude_mask].reset_index(drop=True)

    return filtered_df, excluded_df


def process_data_final(excel_file, sheet_name, header_row):
    """최종 데이터 처리"""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row, dtype=str)
        df.columns = [str(c).upper().strip() for c in df.columns]
        
        rename_map = {
            'VAR NAME': 'ITEM ID', 'VARIABLE NAME': 'ITEM ID', 'VARIABLE': 'ITEM ID', 'OID': 'ITEM ID', 'ITEMOID': 'ITEM ID',
            'FORM': 'PAGE', 'FORM OID': 'PAGE', 'FORM NAME': 'PAGE', 'CRF PAGE': 'PAGE',
            'FOLDER': 'VISIT', 'FOLDER OID': 'VISIT', 'EVENT': 'VISIT',
            'DATASET': 'DOMAIN', 'LB DOMAIN': 'DOMAIN',
            'VER.': 'VERSION', 'VER': 'VERSION', 'CRF_VERSION': 'VERSION', 'CRF VERSION': 'VERSION'
        }
        df = df.rename(columns=rename_map)
        
        std_cols = ['DOMAIN', 'DOMAIN LABEL', 'PAGE', 'PAGE LABEL', 'VISIT', 
                    'ITEM ID', 'ITEM LABEL', 'ITEM SEQ', 'VERSION', 'CODE', 
                    'LAYOUT', 'TYPE', 'MAX_LEN', 'MIN_VAL', 'MAX_VAL']
        
        for col in std_cols:
            if col not in df.columns: df[col] = ""
            df[col] = df[col].fillna("").astype(str).apply(lambda x: x.replace('.0', '').strip() if x.endswith('.0') else x.strip())

        # JOIN KEY 생성
        df['JOIN_KEY'] = (df['DOMAIN'] + df['PAGE'] + df['VISIT'] + df['ITEM ID']).str.replace(r'\s+', '', regex=True).str.upper()
        
        df = df[df['JOIN_KEY'].str.len() > 1]
        df = df.drop_duplicates(subset=['JOIN_KEY'])
        
        return df
    except Exception as e:
        return pd.DataFrame()

def save_to_template(template_path, df_doc, df_edc, ver_info):
    if not os.path.exists(template_path): return None
    wb = load_workbook(template_path)

    target_sheet = 'Entry Screen Validation'
    if target_sheet not in wb.sheetnames: return None
    ws = wb[target_sheet]

    # --- 템플릿 컬럼 위치 동적 파악 (6행 기준) ---
    template_header_row = 6
    doc_col_map = {}
    edc_col_map = {}
    
    for col_idx in range(1, 31):
        col_name = ws.cell(row=template_header_row, column=col_idx).value
        if col_name:
            col_name = str(col_name).strip().upper()
            if col_idx <= 15:
                doc_col_map[col_name] = col_idx
            else:
                edc_col_map[col_name] = col_idx
    
    res_col_idx = 31 
    for col_idx in range(31, ws.max_column + 1):
        if "확인 결과" in str(ws.cell(row=5, column=col_idx).value or "") or \
           "확인 결과" in str(ws.cell(row=6, column=col_idx).value or ""):
            res_col_idx = col_idx
            break

    # --- 데이터 비교 및 입력 ---
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    df_doc['ORIGINAL_ORDER'] = range(len(df_doc))
    merged = pd.merge(df_doc, df_edc, on='JOIN_KEY', how='outer', suffixes=('_Doc', '_EDC'), indicator=True)
    merged = merged.sort_values(by=['ORIGINAL_ORDER'], na_position='last').drop(columns=['ORIGINAL_ORDER'])

    start_row = 7
    for i, row in merged.reset_index(drop=True).iterrows():
        curr_r = start_row + i
        status = row['_merge']
        
        cols_to_fill = list(doc_col_map.keys())
        
        mismatches = []
        if status == 'both':
            for cname in cols_to_fill:
                d_val = str(row.get(f"{cname}_Doc", "")).strip()
                e_val = str(row.get(f"{cname}_EDC", "")).strip()
                if d_val != e_val:
                    mismatches.append(cname)

        for cname, col_idx in doc_col_map.items():
            cell = ws.cell(row=curr_r, column=col_idx)
            val = row.get(f"{cname}_Doc", "") if status != 'right_only' else ""
            cell.value = val
            cell.border = thin_border
            cell.alignment = align_center
            if (status == 'left_only') or (status == 'both' and cname in mismatches):
                cell.fill = red_fill

        for cname, col_idx in edc_col_map.items():
            cell = ws.cell(row=curr_r, column=col_idx)
            val = row.get(f"{cname}_EDC", "") if status != 'left_only' else ""
            cell.value = val
            cell.border = thin_border
            cell.alignment = align_center
            if (status == 'right_only') or (status == 'both' and cname in mismatches):
                cell.fill = red_fill

        res_text = "True" if status == 'both' and not mismatches else "False"
        cell_res = ws.cell(row=curr_r, column=res_col_idx)
        cell_res.value = res_text
        cell_res.border = thin_border
        cell_res.alignment = align_center

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- 3. UI 구성 ---
col1, col2 = st.columns([4, 15], vertical_alignment="center")
logo_path = "JNPMEDI_original.jpg"

with col1:
    st.image(logo_path, width=200) 

with col2:
    st.title("EDC Validation")

st.info("실시간 프리뷰를 통해 컬럼이 올바르게 인식되는지 확인 후 검증을 시작하세요.")

col_u1, col_u2 = st.columns(2)
with col_u1:
    doc_file_up = st.file_uploader("📂 기준 문서 (DB Spec)", type=['xlsx', 'xls'], key="doc")
with col_u2:
    edc_file_up = st.file_uploader("📂 비교 대상 (CDMS Export)", type=['xlsx', 'xls'], key="edc")

if doc_file_up and edc_file_up:
    st.markdown("---")
    
    try:
        doc_excel = load_excel_file(doc_file_up)
        edc_excel = load_excel_file(edc_file_up)
    except Exception as e:
        st.error(f"엑셀 파일 로드 중 오류: {e}")
        st.stop()

    c1, c2 = st.columns(2)

    with c1:
        st.subheader("📄 DB Spec 설정")
        doc_sheet = st.selectbox("시트 선택", doc_excel.sheet_names, key="s1")
        doc_header = st.number_input("헤더 행 (Row Index)", min_value=0, value=1, step=1, key="h1")
        
        doc_df = get_dynamic_preview(doc_excel, doc_sheet, doc_header)
        st.caption(f"▼ '{doc_sheet}' 시트의 {doc_header}번 행을 헤더로 인식한 결과:")
        st.dataframe(doc_df.head(3), use_container_width=True, hide_index=True)
        
        is_ok, msg, missing = check_columns_status(doc_df)
        if is_ok:
            st.markdown(f'<div class="success-box">{msg}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="error-box">{msg}</div>', unsafe_allow_html=True)
        doc_ready = is_ok

    with c2:
        st.subheader("📄 EDC Export 설정")
        edc_sheet = st.selectbox("시트 선택", edc_excel.sheet_names, key="s2")
        edc_header = st.number_input("헤더 행 (Row Index)", min_value=0, value=0, step=1, key="h2")
        
        edc_df = get_dynamic_preview(edc_excel, edc_sheet, edc_header)
        st.caption(f"▼ '{edc_sheet}' 시트의 {edc_header}번 행을 헤더로 인식한 결과:")
        st.dataframe(edc_df.head(3), use_container_width=True, hide_index=True)
        
        is_ok, msg, missing = check_columns_status(edc_df)
        if is_ok:
            st.markdown(f'<div class="success-box">{msg}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="error-box">{msg}</div>', unsafe_allow_html=True)
        edc_ready = is_ok

    st.markdown("---")
    
    with st.expander("📌 버전 정보 (Optional)", expanded=False):
        v1, v2, v3 = st.columns(3)
        bv = v1.text_input("Blank Ver.", "1.0")
        dv = v2.text_input("DB Spec Ver.", "1.0")
        av = v3.text_input("Annotated Ver.", "1.0")

    if not os.path.exists(TEMPLATE_PATH):
        st.error(f"🚨 중요: 실행 경로에 '{TEMPLATE_PATH}' 파일이 없습니다.")
        btn_disabled = True
    else:
        btn_disabled = not (doc_ready and edc_ready)

    if st.button("🚀 검증 시작 (Start Validation)", type="primary", disabled=btn_disabled):
        with st.status("데이터 분석 중...", expanded=True) as status:
            df_final_doc = process_data_final(doc_excel, doc_sheet, doc_header)
            df_final_edc = process_data_final(edc_excel, edc_sheet, edc_header)
            
            if df_final_doc.empty or df_final_edc.empty:
                status.update(label="❌ 오류 발생", state="error")
                st.error("데이터 로드 실패.")
            else:
                # ============================================================
                # [신규] DB Spec에 SYS_ 레이아웃 필터링 적용
                # ============================================================
                df_final_doc, df_excluded = apply_sys_layout_filter(
                    df_final_doc, SYS_LAYOUT_WHITELIST
                )

                # 제외된 항목이 있으면 사이드 정보로 표시
                if not df_excluded.empty:
                    st.info(
                        f"ℹ️ SYS_ 레이아웃으로 인해 비교에서 제외된 항목: "
                        f"**{len(df_excluded)}건** "
                        f"(Whitelist 적용 항목은 비교에 포함됨)"
                    )
                    with st.expander("제외된 항목 확인 (SYS_ 필터)"):
                        st.dataframe(
                            df_excluded[['DOMAIN', 'PAGE', 'VISIT', 'ITEM ID', 'LAYOUT']],
                            use_container_width=True,
                            hide_index=True
                        )
                # ============================================================

                ver_info = {'blank': bv, 'db': dv, 'annotated': av}
                result_file = save_to_template(TEMPLATE_PATH, df_final_doc, df_final_edc, ver_info)
                
                if result_file:
                    status.update(label="🎉 완료!", state="complete")
                    st.success("검증이 성공적으로 완료되었습니다.")
                    
                    today_str = pd.Timestamp.now().strftime('%Y%m%d')
                    file_name = f"EDC Validation List_{today_str}.xlsx"
                    st.download_button(
                        label="📥 결과 리포트 다운로드",
                        data=result_file,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    status.update(label="❌ 템플릿 저장 실패", state="error")
                    st.error("결과 파일 생성 중 오류가 발생했습니다.")
else:
    st.info("👆 먼저 상단에서 두 개의 파일을 업로드해주세요.")